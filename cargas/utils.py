import pandas as pd
import numpy as np
import os
import datetime
import gspread
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from google.oauth2 import service_account
from dotenv import load_dotenv
from datetime import datetime, date
import time
import qrcode
from io import BytesIO
from pathlib import Path
import environ
from typing import Optional
import redis, json, uuid
# import win32print

from django.core.files import File
from django.urls import reverse
from django.db.models import Max
from django.utils.timezone import now
from django.db import transaction
from django.conf import settings
from django.contrib.staticfiles import finders
from django.db.models import Max, F, Q

from apontamento_montagem.models import PecasOrdem as POM
from apontamento_pintura.models import PecasOrdem as POP
from apontamento_solda.models import PecasOrdem as POS
from core.models import Ordem
from cadastro.models import Maquina
from apontamento_exped.utils import chamar_impressora_pecas_montagem, chamar_impressora_pecas_montagem_2

import warnings
warnings.filterwarnings("ignore")

REDIS_URL = "redis://default:AWbmAbD4G2CfZPb3RxwuWQ4RfY7JOmxS@redis-19210.c262.us-east-1-3.ec2.redns.redis-cloud.com:19210"
QUEUE_NAME = "print-zebra"

# Carregar variáveis do arquivo .env
load_dotenv()

BASE_DIR = Path(__file__).resolve().parent.parent

env = environ.Env()
environ.Env.read_env(os.path.join(BASE_DIR, '.env'))

google_credentials_json={
            "type":os.environ.get('type'),
            "project_id":os.environ.get('project_id'),
            "private_key":os.environ.get('private_key'),
            "private_key_id":os.environ.get('private_key_id'),
            "client_x509_cert_url":os.environ.get('client_x509_cert_url'),
            "client_email":os.environ.get('client_email'),
            "auth_uri":os.environ.get('auth_uri'),
            "auth_provider_x509_cert_url":os.environ.get('auth_provider_x509_cert_url'),
            "universe_domain":os.environ.get('universe_domain'),
            "client_id":os.environ.get('client_id'),
            "token_uri":os.environ.get('token_uri'),
        }

if "\\n" in google_credentials_json["private_key"]:
    google_credentials_json["private_key"] = google_credentials_json["private_key"].replace("\\n", "\n")

scope = ['https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive"]

# google_credentials_json["private_key"] = google_credentials_json["private_key"].replace("\\n", "\n")
# google_credentials_json["private_key"] = google_credentials_json["private_key"]#.replace(" ", "\n")

credentials = service_account.Credentials.from_service_account_info(google_credentials_json, scopes=scope)

sa = gspread.authorize(credentials)

name_sheet = 'Bases para sequenciamento'

worksheet1 = 'Base_Carretas'
worksheet2 = 'Carga_Vendas'

sh = sa.open(name_sheet)

def get_data_from_sheets():
    """Carrega os dados das planilhas e retorna como DataFrames."""
    wks1 = sh.worksheet(worksheet1)
    wks2 = sh.worksheet(worksheet2)

    list1 = wks1.get_all_records()
    list2 = wks2.get_all_records()

    # Transformando em dataframes
    base_carretas = pd.DataFrame(list1)
    base_carga = pd.DataFrame(list2)

    return base_carretas, base_carga

def get_base_carreta(carretas: Optional[list] = None):

    """Carrega os dados das planilhas e retorna como DataFrames."""
    wks1 = sh.worksheet(worksheet1)

    list1 = wks1.get_all_records()

    # Transformando em dataframes
    base_carretas = pd.DataFrame(list1)

    if carretas:
        base_carretas = base_carretas[base_carretas['Recurso'].isin(carretas)]

    return base_carretas

def buscar_celulas(carretas):

    base_carretas = get_base_carreta(carretas)
    base_carretas[base_carretas['Etapa2'] == 'Pintura']

    return base_carretas['Célula'].unique()

def tratando_dados(base_carretas, base_carga, data_carga, cliente=None, carga=None):

    ##### Tratando datas######

    # base_carretas, base_carga = get_data_from_sheets()

    dados_carga = base_carga
    dados_carreta = base_carretas

    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].apply(lambda x: "0" + str(x) if len(str(x))==5 else str(x))
    dados_carreta['Recurso'] = dados_carreta['Recurso'].apply(lambda x: "0" + str(x) if len(str(x))==5 else str(x))

    base_carga = base_carga[['UF','Carga','Cidade','PED_PESSOA.CODIGO','PED_NUMEROSERIE','PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO', 'PED_QUANTIDADE']]
    base_carga['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(
        base_carga['PED_PREVISAOEMISSAODOC'], format='%d/%m/%Y', errors='coerce')
    
    # filtrando
    base_carga = base_carga[base_carga['PED_PREVISAOEMISSAODOC'] == data_carga]
    if carga:
        base_carga = base_carga[base_carga['Carga'] == carga]

    if cliente:
        base_carga = base_carga[base_carga['PED_PESSOA.CODIGO'] == cliente]

    # base_carga['Ano'] = base_carga['PED_PREVISAOEMISSAODOC'].dt.strftime('%Y')
    base_carga['PED_PREVISAOEMISSAODOC'] = base_carga.PED_PREVISAOEMISSAODOC.dt.strftime(
        '%d/%m/%Y')

    #### renomeando colunas#####
    base_carga = base_carga.rename(columns={'PED_PREVISAOEMISSAODOC': 'Datas',
                                            'PED_RECURSO.CODIGO': 'Recurso',
                                            'PED_QUANTIDADE': 'Qtde'})

    ##### Valores nulos######

    base_carga.dropna(inplace=True)
    base_carga.reset_index(drop=True)

    # base_carga[base_carga['Recurso'] == '034830CO']

    today = datetime.now()
    ts = pd.Timestamp(today)
    today = today.strftime('%d/%m/%Y')

    return dados_carreta,base_carga

def consultar_carretas(data_inicial, data_final):
    dados_carreta, dados_carga = get_data_from_sheets()
    dados_carreta['Recurso'] = dados_carreta['Recurso'].astype(str)

    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('AM', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('AN', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('VJ', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('LC', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('CE', '') # cinza escuro
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('VM', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('AV', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('CO', '')
    dados_carreta['Recurso'] = dados_carreta['Recurso'].str.replace('SA', '')

    dados_carreta['Recurso'] = dados_carreta['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else str(x))
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else str(x))

    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('AM', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('AN', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('VJ', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('LC', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('CE', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('VM', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('AV', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('CO', '')
    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].str.replace('SA', '')

    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].apply(
        lambda x: x.rstrip()
    )

    dados_carga['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(
        dados_carga['PED_PREVISAOEMISSAODOC'], 
        format="%d/%m/%Y",
        errors='coerce'
    )

    # Filtra pelo intervalo de datas
    dados_carga_data_filtrada = dados_carga[
        (dados_carga['PED_PREVISAOEMISSAODOC'] >= data_inicial) &
        (dados_carga['PED_PREVISAOEMISSAODOC'] <= data_final)
    ]

    dados_carga_data_filtrada['PED_QUANTIDADE'] = dados_carga_data_filtrada['PED_QUANTIDADE'].astype(float)

    # Agrupa os dados por data e código do recurso
    carretas_unica = dados_carga_data_filtrada[['PED_PREVISAOEMISSAODOC', 'PED_RECURSO.CODIGO', 'PED_QUANTIDADE', 'Carga']]
    agrupado = carretas_unica.groupby(['PED_PREVISAOEMISSAODOC', 'PED_RECURSO.CODIGO', 'Carga'])['PED_QUANTIDADE'].sum().reset_index()

    # Ajusta os códigos dos recursos
    dados_carreta['Recurso'] = dados_carreta['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else x)

    agrupado['Contém'] = agrupado['PED_RECURSO.CODIGO'].apply(
        lambda x: '✅' if x in dados_carreta['Recurso'].astype(str).values else '❌'
    )

    buscar_cel = buscar_celulas(
        agrupado['PED_RECURSO.CODIGO'].unique().tolist()
    )

    resultado = [
        {
            "data_carga": str(row['PED_PREVISAOEMISSAODOC'].date()),  # Convertendo para string
            "codigo_recurso": "0" + str(row['PED_RECURSO.CODIGO']) if len(str(row['PED_RECURSO.CODIGO'])) == 5 else str(row['PED_RECURSO.CODIGO']),
            "quantidade": float(row['PED_QUANTIDADE']),
            "presente_no_carreta": row['Contém'],
            "carga": row['Carga']
        }
        for _, row in agrupado.iterrows()
    ]

    celulas = [{"celula": cel} for cel in buscar_cel]

    return {
        "cargas": resultado,
        "celulas": celulas,
    }

def criar_array_datas(data_inicial, data_final):
    # Converte as strings de data para objetos datetime
    data_inicial = datetime.strptime(data_inicial, "%Y-%m-%d")  # Corrigido formato
    data_final = datetime.strptime(data_final, "%Y-%m-%d")  # Corrigido formato
    
    # Gera a lista de datas
    array_datas = []
    data_atual = data_inicial
    
    while data_atual <= data_final:
        # Adiciona a data atual formatada como string à lista
        array_datas.append(data_atual.strftime("%d/%m/%Y"))
        # Avança um dia
        data_atual += timedelta(days=1)
    
    return array_datas

def gerar_arquivos(data_inicial, data_final, setor):
    filenames = []

    resultado = criar_array_datas(data_inicial, data_final)
    base_carretas_original, base_carga_original = get_data_from_sheets()

    resultado = pd.to_datetime(resultado, dayfirst=True, errors='coerce')

    base_carretas_original['Recurso'] = base_carretas_original['Recurso'].astype(str)

    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('AM', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('AN', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('VJ', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('LC', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('VM', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('AV', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('CO', '')

    base_carretas_original['Recurso'] = base_carretas_original['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else str(x))
    base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else str(x))

    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('AM', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('AN', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('VJ', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('LC', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('VM', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('AV', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('CO', '')

    base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].apply(
        lambda x: x.rstrip()
    )

    # Ajusta colunas
    base_carga_original = base_carga_original[['PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO', 'PED_QUANTIDADE']]

    # Converte para datetime
    base_carga_original['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(
        base_carga_original['PED_PREVISAOEMISSAODOC'], format='%d/%m/%Y', dayfirst=True, errors='coerce'
    )

    # Pega apenas o ano ANTES de formatar para string
    # base_carga_original['Ano'] = base_carga_original['PED_PREVISAOEMISSAODOC'].dt.year.astype(str)

    # Formata final para o formato desejado
    # base_carga_original['PED_PREVISAOEMISSAODOC'] = base_carga_original['PED_PREVISAOEMISSAODOC'].dt.strftime('%d/%m/%Y')

    # Renomeia
    base_carga_original = base_carga_original.rename(columns={
        'PED_PREVISAOEMISSAODOC': 'Datas',
        'PED_RECURSO.CODIGO': 'Recurso',
        'PED_QUANTIDADE': 'Qtde'
    })

    base_carga_original.dropna(inplace=True)
    base_carga_original.reset_index(drop=True)

    # base_carga_original[base_carga_original['Recurso'] == 'CHASSI F6 CS M17']
    # base_carretas_original[base_carretas_original['Recurso'] == 'CHASSI F6 CS M17']

    for idx, data_escolhida in enumerate(resultado):
        data_nome_planilha = str(data_escolhida.date()).replace("/","-")
        base_carretas = base_carretas_original.copy()
        base_carga = base_carga_original.copy()

        if setor == 'pintura':

            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)
            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')

            base_carretas.drop(['Etapa', 'Etapa3', 'Etapa4'], axis=1, inplace=True)

            base_carretas.drop(
                base_carretas[(base_carretas['Etapa2'] == '')].index, inplace=True)

            base_carretas = base_carretas.reset_index(drop=True)

            base_carretas = base_carretas.astype(str)
            
            for d in range(0, base_carretas.shape[0]):

                if len(base_carretas['Código'][d]) == 5:
                    base_carretas['Código'][d] = '0' + base_carretas['Código'][d]

            # separando string por "-" e adicionando no dataframe antigo
            base_carga["Recurso"] = base_carga["Recurso"].astype(str)

            tratando_coluna = base_carga["Recurso"].str.split(
                " - ", n=1, expand=True)

            base_carga['Recurso'] = tratando_coluna[0]

            # tratando cores da string

            base_carga['Recurso_cor'] = base_carga['Recurso']

            base_carga = base_carga.reset_index(drop=True)

            df_cores = pd.DataFrame({'Recurso_cor': ['AN', 'VJ', 'LJ', 'LC', 'VM', 'AV', 'sem_cor', 'CO', 'CE'],
                                    'cor': ['Azul', 'Verde', 'Laranja Jacto', 'Laranja', 'Vermelho', 'Amarelo', 'Laranja', 'Cinza', 'Cinza escuro']})

            nome_cor_para_sigla = dict(zip(df_cores['cor'], df_cores['Recurso_cor']))

            cores = ['AM', 'AN', 'VJ', 'LJ', 'LC', 'VM', 'AV', 'CO', 'CE']

            base_carga = base_carga.astype(str)

            for r in range(0, base_carga.shape[0]):
                base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][len(
                    base_carga['Recurso_cor'][r])-3:len(base_carga['Recurso_cor'][r])]
                base_carga['Recurso_cor'] = base_carga['Recurso_cor'].str.strip()

                if len(base_carga['Recurso_cor'][r]) > 2:
                    base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][1:3]

                if base_carga['Recurso_cor'][r] not in cores:
                    base_carga['Recurso_cor'][r] = "LC"

            base_carga = pd.merge(base_carga, df_cores, on=[
                                'Recurso_cor'], how='left')

            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AN', '')  # Azul
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'VJ', '')  # Verde
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'LC', '')  # Laranja
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'VM', '')  # Vermelho
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AV', '')  # Amarelo
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'CO', '')  # Cinza
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'SA', '')  # SEM ADESIVO
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'CE', '')  # Cinza escuro

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == str(data_escolhida.date()))
            filtro_data = base_carga.loc[escolha_data]
            print(filtro_data)

            # filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            # procv e trazendo as colunas que quero ver

            filtro_data = filtro_data.reset_index(drop=True)

            # for i in range(len(filtro_data)):
            #     if filtro_data['Recurso'][i][0] == '0':
            #         filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]

            # tab_completa['Recurso'] = tab_completa['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else x)
            filtro_data['Recurso'] = filtro_data['Recurso'].apply(lambda x: "0" + str(x)  if len(str(x)) == 5 else x)

            tab_completa = pd.merge(filtro_data, base_carretas, on=[
                                    'Recurso'], how='left')

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa = tab_completa.reset_index(drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(drop=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # tratando coluna de código

            for t in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][t]) == 5:
                    tab_completa['Código'][t] = '0' + \
                        tab_completa['Código'][t][0:5]

                if len(tab_completa['Código'][t]) == 8:
                    tab_completa['Código'][t] = tab_completa['Código'][t][0:6]

            # criando coluna de quantidade total de itens

            tab_completa = tab_completa.dropna()

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',', '.')

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa = tab_completa.dropna(axis=0)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']
            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y', 'LEAD TIME', 'flag peça', 'Etapa2'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor']).sum()
            tab_completa.reset_index(inplace=True)

            # linha abaixo exclui eixo simples do sequenciamento da pintura
            # tab_completa.drop(tab_completa.loc[tab_completa['Célula']=='EIXO SIMPLES'].index, inplace=True)
            tab_completa.reset_index(inplace=True, drop=True)

            # Normaliza os valores da coluna 'Etapa5' para identificar corretamente as cores
            tab_completa.loc[tab_completa['Etapa5'].str.contains('CINZA', na=False), 'Etapa5'] = 'CINZA'
            tab_completa.loc[tab_completa['Etapa5'].str.contains('COLORIDO', na=False), 'Etapa5'] = 'COLORIDO'
            tab_completa.loc[tab_completa['Etapa5'].str.contains('PRETO', na=False), 'Etapa5'] = 'PRETO'

            # Função para definir a coluna 'Recurso_cor'
            def definir_recurso_cor(row, corPlanilha, siglaCor):
                if row['Etapa5'] == corPlanilha:
                    return row['Código'] + siglaCor
                else:
                    return row['Código'] + row['Recurso_cor']

            # Função para definir a coluna 'cor'
            def definir_cor(row, corPlanilha, returnCor):
                if row['Etapa5'] == corPlanilha:
                    return returnCor
                else:
                    return row['cor']

            # Aplicando as funções corretamente
            tab_completa['Recurso_cor'] = tab_completa.apply(lambda row: definir_recurso_cor(row, 'CINZA', 'Cinza'), axis=1)
            tab_completa['cor'] = tab_completa.apply(lambda row: definir_cor(row, 'CINZA', 'Cinza'), axis=1)

            tab_completa['Recurso_cor'] = tab_completa.apply(lambda row: definir_recurso_cor(row, 'PRETO', 'Preto'), axis=1)
            tab_completa['cor'] = tab_completa.apply(lambda row: definir_cor(row, 'PRETO', 'Preto'), axis=1)

            # Consumo de tinta

            # tab_completa = tab_completa.merge(df_consumo_pu[['Codigo item','Consumo Pó (kg)','Consumo PU (L)','Consumo Catalisador (L)']], left_on='Código', right_on='Codigo item', how='left').fillna(0)
            
            # tab_completa['Consumo Pó (kg)'] = tab_completa['Consumo Pó (kg)'] * tab_completa['Qtde_total']
            # tab_completa['Consumo PU (L)'] = tab_completa['Consumo PU (L)'] * tab_completa['Qtde_total']
            # tab_completa['Consumo Catalisador (L)'] = tab_completa['Consumo Catalisador (L)'] * tab_completa['Qtde_total']

            # consumo_po = sum(tab_completa['Consumo Pó (kg)'])
            # consumo_po = f'{round(consumo_po / 25, 2)} caixa(s)'

            # consumo_pu_litros = sum(tab_completa['Consumo Pó (kg)'])
            # consumo_pu_latas = round(consumo_pu_litros / 3.08, 2)
            # consumo_pu = f'{consumo_pu_latas} lata(s)'

            # consumo_catalisador_litros = sum(tab_completa['Consumo Catalisador (L)'])
            # consumo_catalisador_latas = round(consumo_catalisador_litros * 1000 / 400, 2)
            # consumo_cata = f'{consumo_catalisador_latas} lata(s)'

            # diluente = f'{round((consumo_pu_litros * 0.80) / 5, 2)} lata(s)'

            ###########################################################################################

            cor_unique = tab_completa['cor'].unique()
            # if idx == 0:
            #     st.write("Arquivos para download")

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()
            file_counter = 1
            rows_per_file = 21
            
            # for carga in carga_unique:
            for i in range(len(cor_unique)):

                start_index = 0
                
                filtro_excel = (tab_completa['cor'] == cor_unique[i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar = filtrar.reset_index(drop=True)
                filtrar = filtrar.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor']
                ).sum().reset_index()
                filtrar.sort_values(by=['Célula'], inplace=True)
                filtrar = filtrar.reset_index(drop=True)

                arquivo_modelo = 'modelo_excel/modelo_op_pintura.xlsx'

                while start_index < len(filtrar):
                # Criar um novo Workbook para cada conjunto de 21 linhas
                # Tente encontrar o arquivo usando finders
                    caminho_modelo = finders.find(arquivo_modelo)
                    print(f"Caminho retornado pelo finder: {caminho_modelo}")
                    
                    # Se não encontrou, vamos verificar o caminho manualmente
                    if not caminho_modelo:
                        # Tente construir caminhos alternativos para localizar o arquivo
                        caminhos_possiveis = [
                            os.path.join(settings.BASE_DIR, 'cargas', 'static', 'modelo_excel', os.path.basename(arquivo_modelo)),
                            os.path.join(settings.STATIC_ROOT, 'modelo_excel', os.path.basename(arquivo_modelo)) if hasattr(settings, 'STATIC_ROOT') else None,
                            os.path.join(settings.BASE_DIR, 'staticfiles', 'modelo_excel', os.path.basename(arquivo_modelo)),
                        ]
                        
                        # Filtra para remover caminhos None
                        caminhos_possiveis = [p for p in caminhos_possiveis if p]
                        
                        # Tente cada um dos caminhos possíveis
                        for caminho in caminhos_possiveis:
                            print(f"Verificando caminho: {caminho}")
                            if os.path.exists(caminho):
                                caminho_modelo = caminho
                                print(f"Arquivo encontrado em: {caminho_modelo}")
                                break
                    
                    # Se ainda não encontrou, levante erro personalizado
                    if not caminho_modelo:
                        raise FileNotFoundError(f"Arquivo {arquivo_modelo} não encontrado em nenhum caminho conhecido")
                    
                    # Carrega o workbook
                    wb = load_workbook(caminho_modelo)
                    ws = wb.active

                    k = 9  # Início da linha no Excel

                    # Define o limite superior para as linhas deste arquivo
                    end_index = min(start_index + rows_per_file, len(filtrar))

                    # Escreve os dados no Excel para as linhas entre start_index e end_index
                    for j in range(start_index, end_index):
                        
                        cor_nome = filtrar['cor'][j]
                        sigla_cor = nome_cor_para_sigla.get(cor_nome, 'sem_cor')

                        ws['F5'] = cor_unique[i]  # nome da coluna é '0'
                        ws['AD5'] = datetime.now()  # data de hoje
                        ws['M4'] = data_escolhida  # data da carga
                        ws['B' + str(k)] = f"{str(filtrar['Código'][j])}{sigla_cor}"
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        ws['K3'] = "N/A"
                        ws['Q3'] = "N/A"
                        ws['AE3'] = "N/A"
                        ws['AN3'] = "N/A"
                        k += 1

                    # Salvar o arquivo com numeração sequencial
                    file_name = f"Pintura {cor_unique[i]} {data_nome_planilha} {file_counter}.xlsx"
                    wb.save(file_name)
                    filenames.append(file_name)
                    
                    # Incrementar índice e contador de arquivos
                    start_index = end_index
                    file_counter += 1

        if setor == 'montagem':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CE', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('SA', '')

            ###### retirando espaco em branco####
            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa3', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa'] == '')].index, inplace=True)
            
            base_carretas = base_carretas.reset_index(drop=True)
            
            for i in range(len(base_carretas)):
                recurso = base_carretas.loc[i, 'Recurso']
                if len(recurso) == 5:
                    base_carretas.loc[i, 'Recurso'] = "0" + recurso

            #### criando código único#####

            codigo_unico = str(data_escolhida.date())[:2] + str(data_escolhida.date())[3:5] + str(data_escolhida.date())[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == str(data_escolhida.date()))
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            filtro_data = filtro_data.reset_index(drop=True)
            # filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            # for i in range(len(filtro_data)):
            #     if filtro_data['Recurso'][i][0] == '0':
            #         filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]
            #     if len(filtro_data['Recurso'][i]) == 5:
            #         filtro_data['Recurso'][i] = "0" + filtro_data['Recurso'][i]
            
            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # base_carretas[base_carretas['Recurso'] == '034538M21']

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = (
                tab_completa['Código']
                    .fillna('')
                    .astype(str)
                    .str.strip()
                    .str.replace(r'\.0$', '', regex=True)
            )

            def _normaliza_codigo(codigo: str) -> str:
                if len(codigo) == 5:
                    return '0' + codigo
                if len(codigo) == 8:
                    return codigo[:6]
                return codigo

            tab_completa['Código'] = tab_completa['Código'].apply(_normaliza_codigo)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa1 = tab_completa[['Código','Peca','Célula','Datas','Carga','Qtde_total']]

            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas','Carga']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + data_escolhida.strftime('%d%m%Y')

            k = 9

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            # print(tab_completa.columns)
            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas', 'Carga', 'PED_CHCRIACAO', 'Ano', 'codigo']).sum()
        
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()

            # for carga in carga_unique:
            file_counter = 1  # Contador de arquivos
            rows_per_file = 21  # Número máximo de linhas por arquivo
            k = 9  # Posição inicial no Excel

            for i in range(0, len(celulas_unique)):
                # Filtrar os dados para a célula atual
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel].reset_index(drop=True)

                # Verificar se o DataFrame está vazio
                if filtrar.empty:
                    continue

                # Índice inicial para controle de divisão em blocos
                start_index = 0

                arquivo_modelo = 'modelo_excel/modelo_op_montagem.xlsx'

                while start_index < len(filtrar):
                    # Criar um novo Workbook para cada conjunto de 21 linhas
                    # Tente encontrar o arquivo usando finders
                    caminho_modelo = finders.find(arquivo_modelo)
                    print(f"Caminho retornado pelo finder: {caminho_modelo}")
                    
                    # Se não encontrou, vamos verificar o caminho manualmente
                    if not caminho_modelo:
                        # Tente construir caminhos alternativos para localizar o arquivo
                        caminhos_possiveis = [
                            os.path.join(settings.BASE_DIR, 'cargas', 'static', 'modelo_excel', os.path.basename(arquivo_modelo)),
                            os.path.join(settings.STATIC_ROOT, 'modelo_excel', os.path.basename(arquivo_modelo)) if hasattr(settings, 'STATIC_ROOT') else None,
                            os.path.join(settings.BASE_DIR, 'staticfiles', 'modelo_excel', os.path.basename(arquivo_modelo)),
                        ]
                        
                        # Filtra para remover caminhos None
                        caminhos_possiveis = [p for p in caminhos_possiveis if p]
                        
                        # Tente cada um dos caminhos possíveis
                        for caminho in caminhos_possiveis:
                            print(f"Verificando caminho: {caminho}")
                            if os.path.exists(caminho):
                                caminho_modelo = caminho
                                print(f"Arquivo encontrado em: {caminho_modelo}")
                                break
                    
                    # Se ainda não encontrou, levante erro personalizado
                    if not caminho_modelo:
                        raise FileNotFoundError(f"Arquivo {arquivo_modelo} não encontrado em nenhum caminho conhecido")
                    
                    # Carrega o workbook
                    wb = load_workbook(caminho_modelo)
                    ws = wb.active

                    # Define o limite superior para as linhas deste arquivo
                    end_index = min(start_index + rows_per_file, len(filtrar))
                    k = 9  # Reseta a linha inicial para cada novo arquivo

                    # Escreve os dados no Excel para as linhas entre start_index e end_index
                    for j in range(start_index, end_index):
                        ws['G5'] = celulas_unique[0][i]  # Nome da célula
                        ws['AD5'] = hoje  # Data de hoje

                        # Gerar código único baseado na célula
                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C"
                        elif celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                        else:
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        # Preenchimento das células do Excel
                        ws['M4'] = data_escolhida  # Data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k += 1

                    # Salvar o arquivo com numeração sequencial
                    file_name = f"Montagem {celulas_unique[0][i]} {data_nome_planilha} {file_counter}.xlsx"
                    wb.template = False
                    wb.save(file_name)
                    filenames.append(file_name)

                    # Incrementar o índice de início e o contador de arquivos
                    start_index = end_index
                    file_counter += 1

        if setor == 'solda':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CE', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('SA', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa3'] == '')].index, inplace=True)
            
            base_carretas = base_carretas.reset_index(drop=True)
            
            for i in range(len(base_carretas)):
                recurso = base_carretas.loc[i, 'Recurso']
                if len(recurso) == 5:
                    base_carretas.loc[i, 'Recurso'] = "0" + recurso

            #### criando código único#####

            codigo_unico = str(data_escolhida.date())[:2] + str(data_escolhida.date())[3:5] + str(data_escolhida.date())[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == str(data_escolhida.date()))
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            filtro_data = filtro_data.reset_index(drop=True)
            filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            for i in range(len(filtro_data)):
                recurso = filtro_data.loc[i, 'Recurso']
                if recurso and recurso[0] == '0':
                    recurso = recurso[1:]
                if len(recurso) == 5:
                    recurso = "0" + recurso
                filtro_data.loc[i, 'Recurso'] = recurso
            
            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # base_carretas[base_carretas['Recurso'] == '034538M21']

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = (
                tab_completa['Código']
                    .fillna('')
                    .astype(str)
                    .str.strip()
                    .str.replace(r'\.0$', '', regex=True)
            )

            def _normaliza_codigo(codigo: str) -> str:
                if len(codigo) == 5:
                    return '0' + codigo
                if len(codigo) == 8:
                    return codigo[:6]
                return codigo

            tab_completa['Código'] = tab_completa['Código'].apply(_normaliza_codigo)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa1 = tab_completa[['Código','Peca','Célula','Datas','Carga','Qtde_total']]

            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas','Carga']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + data_escolhida.strftime('%d%m%Y')

            k = 9

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            # print(tab_completa.columns)
            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas', 'Carga', 'PED_CHCRIACAO', 'Ano', 'codigo']).sum()
        
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()

            # for carga in carga_unique:
            file_counter = 1  # Contador de arquivos
            rows_per_file = 21  # Número máximo de linhas por arquivo
            k = 9  # Posição inicial no Excel

            for i in range(0, len(celulas_unique)):
                # Filtrar os dados para a célula atual
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel].reset_index(drop=True)

                # Verificar se o DataFrame está vazio
                if filtrar.empty:
                    continue

                # Índice inicial para controle de divisão em blocos
                start_index = 0

                arquivo_modelo = 'modelo_excel/modelo_op_solda.xlsx'

                while start_index < len(filtrar):
                    # Criar um novo Workbook para cada conjunto de 21 linhas
                    # Tente encontrar o arquivo usando finders
                    caminho_modelo = finders.find(arquivo_modelo)
                    print(f"Caminho retornado pelo finder: {caminho_modelo}")
                    
                    # Se não encontrou, vamos verificar o caminho manualmente
                    if not caminho_modelo:
                        # Tente construir caminhos alternativos para localizar o arquivo
                        caminhos_possiveis = [
                            os.path.join(settings.BASE_DIR, 'cargas', 'static', 'modelo_excel', os.path.basename(arquivo_modelo)),
                            os.path.join(settings.STATIC_ROOT, 'modelo_excel', os.path.basename(arquivo_modelo)) if hasattr(settings, 'STATIC_ROOT') else None,
                            os.path.join(settings.BASE_DIR, 'staticfiles', 'modelo_excel', os.path.basename(arquivo_modelo)),
                        ]
                        
                        # Filtra para remover caminhos None
                        caminhos_possiveis = [p for p in caminhos_possiveis if p]
                        
                        # Tente cada um dos caminhos possíveis
                        for caminho in caminhos_possiveis:
                            print(f"Verificando caminho: {caminho}")
                            if os.path.exists(caminho):
                                caminho_modelo = caminho
                                print(f"Arquivo encontrado em: {caminho_modelo}")
                                break
                    
                    # Se ainda não encontrou, levante erro personalizado
                    if not caminho_modelo:
                        raise FileNotFoundError(f"Arquivo {arquivo_modelo} não encontrado em nenhum caminho conhecido")
                    
                    # Carrega o workbook
                    wb = load_workbook(caminho_modelo)
                    ws = wb.active

                    # Define o limite superior para as linhas deste arquivo
                    end_index = min(start_index + rows_per_file, len(filtrar))
                    k = 9  # Reseta a linha inicial para cada novo arquivo

                    # Escreve os dados no Excel para as linhas entre start_index e end_index
                    for j in range(start_index, end_index):
                        ws['G5'] = celulas_unique[0][i]  # Nome da célula
                        ws['AD5'] = hoje  # Data de hoje

                        # Gerar código único baseado na célula
                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C"
                        elif celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                        else:
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        # Preenchimento das células do Excel
                        ws['M4'] = data_escolhida  # Data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k += 1

                    # Salvar o arquivo com numeração sequencial
                    file_name = f"Solda {celulas_unique[0][i]} {data_nome_planilha} {file_counter}.xlsx"
                    wb.template = False
                    wb.save(file_name)
                    filenames.append(file_name)

                    # Incrementar o índice de início e o contador de arquivos
                    start_index = end_index
                    file_counter += 1

        tab_completa = pd.concat([tab_completa, tab_completa], ignore_index=True)
    
    return filenames

def gerar_sequenciamento(data_inicial, data_final, setor, carga: Optional[str] = None, celula: Optional[str]= None):
    filenames = []
    def _dbg(step, **kwargs):
        try:
            details = " | ".join([f"{k}={v}" for k, v in kwargs.items()])
            msg = f"[gerar_sequenciamento] {step}"
            if details:
                msg = f"{msg} | {details}"
            print(msg, flush=True)
        except Exception as e:
            print(f"[gerar_sequenciamento] log_error | step={step} | error={repr(e)}", flush=True)

    def _dbg_df(step, df, cols=None):
        info = {"rows": len(df), "cols": len(df.columns)}
        if cols:
            for c in cols:
                if c in df.columns:
                    info[f"null_{c}"] = int(df[c].isna().sum())
                    info[f"uniq_{c}"] = int(df[c].nunique(dropna=True))
        _dbg(step, **info)

    _dbg(
        "start",
        data_inicial=data_inicial,
        data_final=data_final,
        setor=setor,
        carga=carga,
        celula=celula
    )

    resultado = criar_array_datas(data_inicial, data_final)
    base_carretas_original, base_carga_original = get_data_from_sheets()
    _dbg("datas_criadas", total_datas=len(resultado), primeira_data=(resultado[0] if len(resultado) > 0 else None), ultima_data=(resultado[-1] if len(resultado) > 0 else None))
    _dbg_df("base_carretas_original_carregada", base_carretas_original, cols=["Recurso", "Etapa2"])
    _dbg_df("base_carga_original_carregada", base_carga_original, cols=["PED_RECURSO.CODIGO", "PED_PREVISAOEMISSAODOC", "Carga"])

    base_carretas_original['Recurso'] = base_carretas_original['Recurso'].astype(str)

    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('AM', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('AN', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('VJ', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('LC', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('VM', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('AV', '')
    # base_carretas_original['Recurso'] = base_carretas_original['Recurso'].str.replace('CO', '')

    base_carretas_original['Recurso'] = base_carretas_original['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else str(x))
    base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else str(x))

    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('AM', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('AN', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('VJ', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('LC', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('VM', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('AV', '')
    # base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].str.replace('CO', '')

    base_carga_original['PED_RECURSO.CODIGO'] = base_carga_original['PED_RECURSO.CODIGO'].apply(
        lambda x: x.rstrip()
    )

    resultado = pd.to_datetime(resultado, dayfirst=True, errors='coerce')
    _dbg("datas_convertidas_datetime", total_datas=len(resultado), datas_invalidas=int(pd.Series(resultado).isna().sum()))

    # Ajusta colunas
    if carga:
        base_carga_original = base_carga_original[['PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO', 'PED_QUANTIDADE', 'Carga']]
    else:
        base_carga_original = base_carga_original[['PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO', 'PED_QUANTIDADE']]
    _dbg_df("base_carga_original_colunas_ajustadas", base_carga_original, cols=["PED_PREVISAOEMISSAODOC", "PED_RECURSO.CODIGO", "Carga"])
 
    # Apenas para pintura
    if celula and setor == 'pintura':
        _dbg_df("base_carretas_antes_filtro_celula", base_carretas_original, cols=["Recurso", "Etapa2", "Célula"])
        base_carretas_original = base_carretas_original[(base_carretas_original['Célula'] == celula) & (base_carretas_original['Etapa2'] == 'Pintura')]
        _dbg_df("base_carretas_depois_filtro_celula", base_carretas_original, cols=["Recurso", "Etapa2", "Célula"])

    # Converte para datetime
    base_carga_original['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(
        base_carga_original['PED_PREVISAOEMISSAODOC'], format='%d/%m/%Y', dayfirst=True, errors='coerce'
    )
    _dbg("datas_base_carga_convertidas", total=len(base_carga_original), datas_invalidas=int(base_carga_original['PED_PREVISAOEMISSAODOC'].isna().sum()))

    # Pega apenas o ano ANTES de formatar para string
    # base_carga_original['Ano'] = base_carga_original['PED_PREVISAOEMISSAODOC'].dt.year.astype(str)

    # Formata final para o formato desejado
    # base_carga_original['PED_PREVISAOEMISSAODOC'] = base_carga_original['PED_PREVISAOEMISSAODOC'].dt.strftime('%d/%m/%Y')

    # Renomeia
    base_carga_original = base_carga_original.rename(columns={
        'PED_PREVISAOEMISSAODOC': 'Datas',
        'PED_RECURSO.CODIGO': 'Recurso',
        'PED_QUANTIDADE': 'Qtde'
    })

    base_carga_original.dropna(inplace=True)
    base_carga_original.reset_index(drop=True)
    _dbg_df("base_carga_original_pos_dropna", base_carga_original, cols=["Datas", "Recurso", "Qtde", "Carga"])
    _dbg_df("base_carretas_original_pre_loop", base_carretas_original, cols=["Recurso", "Etapa2"])

    tab_resultado = pd.DataFrame() 

    for idx, data_escolhida in enumerate(resultado):
        _dbg("loop_inicio", idx=idx, data_escolhida=data_escolhida, setor=setor)
        # data_nome_planilha = data_escolhida.replace("/","-")[:5]
        base_carretas = base_carretas_original.copy()
        base_carga = base_carga_original.copy()
        _dbg_df("loop_base_carretas_copy", base_carretas, cols=["Recurso", "Etapa2"])
        _dbg_df("loop_base_carga_copy", base_carga, cols=["Datas", "Recurso", "Qtde", "Carga"])

        if setor == 'pintura':
            _dbg("pintura_inicio", idx=idx, data=data_escolhida)

            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)
            # base_carretas[base_carretas['Recurso'] == '034550G']
            base_carga['Recurso'] = base_carga['Recurso'].astype(str)
            # base_carga[base_carga['Recurso'] == '034550G']
            
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            # base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')

            base_carretas.drop(['Etapa', 'Etapa3', 'Etapa4'], axis=1, inplace=True)

            base_carretas.drop(
                base_carretas[(base_carretas['Etapa2'] == '')].index, inplace=True)

            base_carretas = base_carretas.reset_index(drop=True)

            base_carretas = base_carretas.astype(str)
            
            for i in range(len(base_carretas)):
                recurso = base_carretas.loc[i, 'Recurso']
                if len(recurso) == 5:
                    base_carretas.loc[i, 'Recurso'] = "0" + recurso

            # separando string por "-" e adicionando no dataframe antigo
            base_carga["Recurso"] = base_carga["Recurso"].astype(str)

            tratando_coluna = base_carga["Recurso"].str.split(
                " - ", n=1, expand=True)

            base_carga['Recurso'] = tratando_coluna[0]

            # tratando cores da string

            base_carga['Recurso_cor'] = base_carga['Recurso']

            base_carga = base_carga.reset_index(drop=True)

            df_cores = pd.DataFrame({'Recurso_cor': ['AN', 'VJ', 'LJ', 'LC', 'VM', 'AV', 'sem_cor', 'CO', 'CE'],
                                    'cor': ['Azul', 'Verde', 'Laranja Jacto', 'Laranja', 'Vermelho', 'Amarelo', 'Laranja', 'Cinza', 'Cinza escuro']})

            nome_cor_para_sigla = dict(zip(df_cores['cor'], df_cores['Recurso_cor']))

            cores = ['AM', 'AN', 'VJ', 'LJ', 'LC', 'VM', 'AV', 'CO', 'CE']

            base_carga = base_carga.astype(str)

            for r in range(0, base_carga.shape[0]):
                base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][len(
                    base_carga['Recurso_cor'][r])-3:len(base_carga['Recurso_cor'][r])]
                base_carga['Recurso_cor'] = base_carga['Recurso_cor'].str.strip()

                if len(base_carga['Recurso_cor'][r]) > 2:
                    base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][1:3]

                if base_carga['Recurso_cor'][r] not in cores:
                    base_carga['Recurso_cor'][r] = "LC"

            base_carga = pd.merge(base_carga, df_cores, on=[
                                'Recurso_cor'], how='left')

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CE', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('SA', '')

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())
            _dbg("pintura_datas_unique", qtd_datas=len(datas_unique))

            escolha_data = (base_carga['Datas'] == str(data_escolhida.date()))
            filtro_data = base_carga.loc[escolha_data]
            _dbg_df("pintura_filtro_data", filtro_data, cols=["Datas", "Recurso", "Qtde", "Carga"])
            # filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            if carga:
                escolha_carga = (base_carga['Carga'] == carga)
                filtro_data = filtro_data.loc[escolha_carga]
                _dbg_df("pintura_filtro_carga", filtro_data, cols=["Datas", "Recurso", "Qtde", "Carga"])

            # procv e trazendo as colunas que quero ver

            filtro_data = filtro_data.reset_index(drop=True)

            # for i in range(len(filtro_data)):
            #     if filtro_data['Recurso'][i][0] == '0':
            #         filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]

            # tab_completa['Recurso'] = tab_completa['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else x)
            filtro_data['Recurso'] = filtro_data['Recurso'].apply(lambda x: "0" + str(x)  if len(str(x)) == 5 else x)
            _dbg("pintura_chave_merge_recurso", recurso_filtro_unique=int(filtro_data['Recurso'].nunique(dropna=True)), recurso_carretas_unique=int(base_carretas['Recurso'].nunique(dropna=True)))

            tab_completa = pd.merge(filtro_data, base_carretas, on=[
                                    'Recurso'], how='left')
            _dbg_df("pintura_tab_completa_pos_merge", tab_completa, cols=["Recurso", "Qtde_x", "Qtde_y"])

            # tab_completa['Código'] = tab_completa['Código'].astype(str)

            # tab_completa['Código'] = (
            #     tab_completa['Código']
            #         .fillna('')                 # evita NaN
            #         .astype(str)
            #         .str.strip()
            #         .str.replace(r'\.0$', '', regex=True)  # remove ".0" no final (ex: 12345.0)
            # )

            tab_completa = tab_completa.reset_index(drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(drop=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # tratando coluna de código

            # for t in range(tab_completa.shape[0]):
            #     codigo = tab_completa['Código'].iloc[t]

            #     if len(codigo) == 5:
            #         tab_completa.loc[t, 'Código'] = '0' + codigo[:5]

            #     elif len(codigo) == 8:
            #         tab_completa.loc[t, 'Código'] = codigo[:6]

            # criando coluna de quantidade total de itens

            tab_completa['Qtde_x'] = pd.to_numeric(
                tab_completa['Qtde_x'].astype(str).str.replace(',', '.'),
                errors='coerce'
            )
            tab_completa['Qtde_y'] = pd.to_numeric(
                tab_completa['Qtde_y'],
                errors='coerce'
            )
            tab_completa = tab_completa.dropna(subset=['Qtde_x', 'Qtde_y'])
            _dbg_df("pintura_tab_completa_pos_dropna_qtde", tab_completa, cols=["Recurso", "Qtde_x", "Qtde_y"])

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']
            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y', 'LEAD TIME', 'flag peça', 'Etapa2'])

            if carga:
                tab_completa = tab_completa.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor', 'Carga']).sum()
            else:
                tab_completa = tab_completa.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor']).sum()

            tab_completa.reset_index(inplace=True)
            _dbg_df("pintura_tab_completa_pos_groupby", tab_completa, cols=["Datas", "Recurso_cor", "cor"])

            # linha abaixo exclui eixo simples do sequenciamento da pintura
            # tab_completa.drop(tab_completa.loc[tab_completa['Célula']=='EIXO SIMPLES'].index, inplace=True)
            tab_completa.reset_index(inplace=True, drop=True)

            tab_completa['Etapa5'].unique()

            # Normaliza os valores da coluna 'Etapa5' para identificar corretamente as cores
            tab_completa.loc[tab_completa['Etapa5'].str.contains('CINZA', na=False), 'Etapa5'] = 'CINZA'
            tab_completa.loc[tab_completa['Etapa5'].str.contains('COLORIDO', na=False), 'Etapa5'] = 'COLORIDO'
            tab_completa.loc[tab_completa['Etapa5'].str.contains('PRETO', na=False), 'Etapa5'] = 'PRETO'

            # Função para definir a coluna 'Recurso_cor'
            def definir_recurso_cor(row, corPlanilha, siglaCor):
                if row['Etapa5'] == corPlanilha:
                    return row['Código'] + siglaCor
                else:
                    return row['Código'] + row['Recurso_cor']

            # Função para definir a coluna 'cor'
            def definir_cor(row, corPlanilha, returnCor):
                if row['Etapa5'] == corPlanilha:
                    return returnCor
                else:
                    return row['cor']

            # Aplicando as funções corretamente
            tab_completa['Recurso_cor'] = tab_completa.apply(lambda row: definir_recurso_cor(row, 'CINZA', 'Cinza'), axis=1)
            tab_completa['cor'] = tab_completa.apply(lambda row: definir_cor(row, 'CINZA', 'Cinza'), axis=1)

            tab_completa['Recurso_cor'] = tab_completa.apply(lambda row: definir_recurso_cor(row, 'PRETO', 'Preto'), axis=1)
            tab_completa['cor'] = tab_completa.apply(lambda row: definir_cor(row, 'PRETO', 'Preto'), axis=1)
            _dbg_df("pintura_tab_completa_final_bloco", tab_completa, cols=["Datas", "Recurso_cor", "cor", "Qtde_total"])

            

            ###########################################################################################

            # if idx == 0:
            #     st.write("Arquivos para download")

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            tab_completa = tab_completa.reset_index(drop=True)

        if setor == 'montagem':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CE', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('SA', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa3', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa'] == '')].index, inplace=True)
            
            base_carretas = base_carretas.reset_index(drop=True)
            
            for i in range(len(base_carretas)):
                recurso = base_carretas.loc[i, 'Recurso']
                if len(recurso) == 5:
                    base_carretas.loc[i, 'Recurso'] = "0" + recurso

            #### criando código único#####

            # codigo_unico = data_escolhida[:2] + data_escolhida[3:5] + data_escolhida[6:10]

            #### filtrando data da carga#####

            # datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == str(data_escolhida.date()))
            filtro_data = base_carga.loc[escolha_data]
            # filtro_data[filtro_data['Recurso'] == '034550G']

            filtro_data = filtro_data.reset_index(drop=True)
            filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            for i in range(len(filtro_data)):
                recurso = filtro_data.loc[i, 'Recurso']
                if recurso and recurso[0] == '0':
                    recurso = recurso[1:]
                if len(recurso) == 5:
                    recurso = "0" + recurso
                filtro_data.loc[i, 'Recurso'] = recurso
            
            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # tab_completa[tab_completa['Recurso'] == '034550G']

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = (
                tab_completa['Código']
                    .fillna('')
                    .astype(str)
                    .str.strip()
                    .str.replace(r'\.0$', '', regex=True)
            )

            def _normaliza_codigo(codigo: str) -> str:
                if len(codigo) == 5:
                    return '0' + codigo
                if len(codigo) == 8:
                    return codigo[:6]
                return codigo

            tab_completa['Código'] = tab_completa['Código'].apply(_normaliza_codigo)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            if carga:
                tab_completa = tab_completa.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas', 'Carga']).sum()
            else:
                tab_completa = tab_completa.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa1 = tab_completa[['Código','Peca','Célula','Datas','Carga','Qtde_total']]

            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas','Carga']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + data_escolhida.strftime('%d%m%Y')

            k = 9

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            # print(tab_completa.columns)
            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas', 'Carga', 'PED_CHCRIACAO', 'Ano', 'codigo']).sum()
        
            tab_completa = tab_completa.reset_index(drop=True)
            # tab_completa[tab_completa['Código'] == '034550'] 
            # carga_unique = tab_completa['Carga'].unique()

            # for carga in carga_unique:
            file_counter = 1  # Contador de arquivos
            rows_per_file = 21  # Número máximo de linhas por arquivo
            k = 9  # Posição inicial no Excel

            # for i in range(0, len(celulas_unique)):
            #     # Filtrar os dados para a célula atual
            #     filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
            #     filtrar = tab_completa.loc[filtro_excel].reset_index(drop=True)

            #     # Verificar se o DataFrame está vazio
            #     if filtrar.empty:
            #         continue

            #     # Índice inicial para controle de divisão em blocos
            #     start_index = 0

            #     while start_index < len(filtrar):
            #         # Criar um novo Workbook para cada conjunto de 21 linhas
            #         wb = Workbook()
            #         wb = load_workbook('modelo_op_montagem.xlsx')
            #         ws = wb.active

            #         # Define o limite superior para as linhas deste arquivo
            #         end_index = min(start_index + rows_per_file, len(filtrar))
            #         k = 9  # Reseta a linha inicial para cada novo arquivo

            #         # Escreve os dados no Excel para as linhas entre start_index e end_index
            #         for j in range(start_index, end_index):
            #             ws['G5'] = celulas_unique[0][i]  # Nome da célula
            #             ws['AD5'] = hoje  # Data de hoje

            #             # Gerar código único baseado na célula
            #             if celulas_unique[0][i] == "EIXO COMPLETO":
            #                 ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C"
            #             elif celulas_unique[0][i] == "EIXO SIMPLES":
            #                 ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
            #             else:
            #                 ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

            #             # Preenchimento das células do Excel
            #             ws['M4'] = data_escolhida  # Data da carga
            #             ws['B' + str(k)] = filtrar['Código'][j]
            #             ws['G' + str(k)] = filtrar['Peca'][j]
            #             ws['AD' + str(k)] = filtrar['Qtde_total'][j]
            #             k += 1

            #         # Salvar o arquivo com numeração sequencial
            #         file_name = f"Montagem {celulas_unique[0][i]} {data_nome_planilha} {file_counter}.xlsx"
            #         wb.template = False
            #         wb.save(file_name)
            #         filenames.append(file_name)

            #         # Incrementar o índice de início e o contador de arquivos
            #         start_index = end_index
            #         file_counter += 1

            # Preparar os dados para inserção no banco de dados
            # data_formatada = datetime.strptime(data_escolhida, '%d/%m/%Y').strftime('%Y-%m-%d')
            # tab_completa['Datas'] = data_formatada
            # data_insert_sql = tab_completa[['Célula', 'Código', 'Peca', 'Qtde_total', 'Datas']].values.tolist()

            # Chamar a função de inserção
            # insert_montagem(data_formatada, data_insert_sql, check_atualizar_base_carga)

        if setor == 'solda':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CE', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('SA', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa3'] == '')].index, inplace=True)
            
            base_carretas = base_carretas.reset_index(drop=True)
            
            for i in range(len(base_carretas)):
                if len(base_carretas['Recurso'][i]) == 5:
                    base_carretas['Recurso'][i] = "0" + base_carretas['Recurso'][i]

            #### criando código único#####

            # codigo_unico = data_escolhida[:2] + data_escolhida[3:5] + data_escolhida[6:10]

            #### filtrando data da carga#####

            # datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == str(data_escolhida.date()))
            filtro_data = base_carga.loc[escolha_data]
            # filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            filtro_data = filtro_data.reset_index(drop=True)
            filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            for i in range(len(filtro_data)):
                if filtro_data['Recurso'][i][0] == '0':
                    filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]
                if len(filtro_data['Recurso'][i]) == 5:
                    filtro_data['Recurso'][i] = "0" + filtro_data['Recurso'][i]
            
            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # base_carretas[base_carretas['Recurso'] == '034538M21']

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            if carga:
                tab_completa = tab_completa.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas', 'Carga']).sum()
            else:
                tab_completa = tab_completa.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa1 = tab_completa[['Código','Peca','Célula','Datas','Carga','Qtde_total']]

            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas','Carga']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            for d in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][d]) == 5:
                    tab_completa['Código'][d] = '0' + tab_completa['Código'][d]

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + data_escolhida.strftime('%d%m%Y')

            k = 9

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            # print(tab_completa.columns)
            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas', 'Carga', 'PED_CHCRIACAO', 'Ano', 'codigo']).sum()
        
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()

            # for carga in carga_unique:
            file_counter = 1  # Contador de arquivos
            rows_per_file = 21  # Número máximo de linhas por arquivo
            k = 9  # Posição inicial no Excel

            # for i in range(0, len(celulas_unique)):
            #     # Filtrar os dados para a célula atual
            #     filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
            #     filtrar = tab_completa.loc[filtro_excel].reset_index(drop=True)

            #     # Verificar se o DataFrame está vazio
            #     if filtrar.empty:
            #         continue

            #     # Índice inicial para controle de divisão em blocos
            #     start_index = 0

            #     while start_index < len(filtrar):
            #         # Criar um novo Workbook para cada conjunto de 21 linhas
            #         wb = Workbook()
            #         wb = load_workbook('modelo_op_montagem.xlsx')
            #         ws = wb.active

            #         # Define o limite superior para as linhas deste arquivo
            #         end_index = min(start_index + rows_per_file, len(filtrar))
            #         k = 9  # Reseta a linha inicial para cada novo arquivo

            #         # Escreve os dados no Excel para as linhas entre start_index e end_index
            #         for j in range(start_index, end_index):
            #             ws['G5'] = celulas_unique[0][i]  # Nome da célula
            #             ws['AD5'] = hoje  # Data de hoje

            #             # Gerar código único baseado na célula
            #             if celulas_unique[0][i] == "EIXO COMPLETO":
            #                 ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C"
            #             elif celulas_unique[0][i] == "EIXO SIMPLES":
            #                 ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
            #             else:
            #                 ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

            #             # Preenchimento das células do Excel
            #             ws['M4'] = data_escolhida  # Data da carga
            #             ws['B' + str(k)] = filtrar['Código'][j]
            #             ws['G' + str(k)] = filtrar['Peca'][j]
            #             ws['AD' + str(k)] = filtrar['Qtde_total'][j]
            #             k += 1

            #         # Salvar o arquivo com numeração sequencial
            #         file_name = f"Montagem {celulas_unique[0][i]} {data_nome_planilha} {file_counter}.xlsx"
            #         wb.template = False
            #         wb.save(file_name)
            #         filenames.append(file_name)

            #         # Incrementar o índice de início e o contador de arquivos
            #         start_index = end_index
            #         file_counter += 1

            # Preparar os dados para inserção no banco de dados
            # data_formatada = datetime.strptime(data_escolhida, '%d/%m/%Y').strftime('%Y-%m-%d')
            # tab_completa['Datas'] = data_formatada
            # data_insert_sql = tab_completa[['Célula', 'Código', 'Peca', 'Qtde_total', 'Datas']].values.tolist()

            # Chamar a função de inserção
            # insert_montagem(data_formatada, data_insert_sql, check_atualizar_base_carga)

        _dbg_df("loop_tab_completa_antes_concat", tab_completa)
        tab_resultado = pd.concat([tab_completa, tab_resultado], ignore_index=True)
        _dbg_df("loop_tab_resultado_pos_concat", tab_resultado)
    
    _dbg_df("fim_tab_resultado", tab_resultado)
    return tab_resultado

def processar_ordens_montagem(request, ordens_data, atualizacao_ordem=None, grupo_maquina='montagem'):

    print(ordens_data)

    if not ordens_data:
        return {"error": "Nenhuma ordem nova adicionada, caso tivesse ordem para ser excluída, foi excluída.", "status": 400}

    # Coletar datas únicas e validar
    try:
        datas_requisicao = set()
        for o in ordens_data:
            data_carga = o.get("data_carga")
            if data_carga:
                if isinstance(data_carga, str):
                    # Se for string, converte usando strptime
                    data_obj = datetime.strptime(data_carga, "%Y-%m-%d").date()
                elif isinstance(data_carga, (datetime, date)):
                    # Se já for datetime ou date, converte para date (se necessário)
                    data_obj = data_carga.date() if isinstance(data_carga, datetime) else data_carga
                else:
                    raise ValueError(f"Tipo inválido de data: {type(data_carga)}")
                datas_requisicao.add(data_obj)
    except Exception as e:
        print(f"Erro ao processar datas: {e}")
        return {"error": "Formato de data inválido! Use YYYY-MM-DD.", "status": 400}

    # Verifica datas já com carga
    datas_existentes = set(
        Ordem.objects.filter(data_carga__in=datas_requisicao, grupo_maquina=grupo_maquina)
        .values_list("data_carga", flat=True)
    )

    datas_bloqueadas = datas_existentes & datas_requisicao
    if not atualizacao_ordem and datas_bloqueadas:
        return {
            "error": f"Datas já com carga alocada: {', '.join(map(str, datas_bloqueadas))}",
            "status": 400
        }

    # Verifica máquinas existentes
    maquinas_requisicao = {o.get("setor_conjunto") for o in ordens_data if o.get("setor_conjunto")}
    maquinas_existentes = set(Maquina.objects.filter(nome__in=maquinas_requisicao).values_list("nome", flat=True))
    maquinas_faltantes = maquinas_requisicao - maquinas_existentes
    if maquinas_faltantes:
        return {
            "error": f"Máquinas não cadastradas: {', '.join(maquinas_faltantes)}",
            "status": 400
        }

    # Coletar datas únicas e validar
    # try:
    #     # formato_data = "%Y-%m-%m" if grupo_maquina == "montagem" else "%Y-%m-%d"
    #     datas_requisicao = {
    #         datetime.strptime(o["data_carga"], "%Y-%m-%d").date()
    #         for o in ordens_data if o.get("data_carga")
    #     }
    # except ValueError:
    #     return {"error": "Formato de data inválido! Use YYYY-MM-DD.", "status": 400}

    # Criação em lote
    # Pega a última ordem atual no banco
    ultimo_numero = Ordem.objects.filter(grupo_maquina=grupo_maquina).aggregate(Max('ordem'))['ordem__max'] or 0

    with transaction.atomic():
        ordens_objs = []
        ordens_metadata = []

        for i, o in enumerate(ordens_data):
            
            # try:
            #     data_carga = datetime.strptime(o["data_carga"], "%Y-%d-%m").date()
            # except:
            # data_carga pode ser str, datetime ou date
            data_carga = o.get("data_carga")

            if isinstance(data_carga, str):
                data_carga = datetime.strptime(data_carga, "%Y-%m-%d").date()
            elif isinstance(data_carga, datetime):
                data_carga = data_carga.date()
            elif isinstance(data_carga, date):
                data_carga = data_carga
            else:
                raise ValueError(f"Tipo inválido para data_carga: {type(data_carga)}")

            nova_ordem = Ordem(
                grupo_maquina=grupo_maquina,
                status_atual="aguardando_iniciar",
                obs=o.get("obs", ""),
                cor=o.get("cor"),
                data_criacao=now(),
                data_carga=data_carga,
                ordem=ultimo_numero + i + 1  # atribui manualmente a ordem
            )

            try:
                maquina = Maquina.objects.get(nome=o["setor_conjunto"],setor__nome='montagem')
                nova_ordem.maquina = maquina
                # calcula data_programacao manualmente
                nova_ordem.data_programacao = data_carga - timedelta(days=3)
                while nova_ordem.data_programacao.weekday() in [5, 6]:
                    nova_ordem.data_programacao -= timedelta(days=1)
            except Maquina.DoesNotExist:
                return {"error": f"Máquina '{o['setor_conjunto']}' não cadastrada.", "status": 400}

            ordens_objs.append(nova_ordem)
            ordens_metadata.append({
                "peca_nome": o["peca_nome"],
                "qtd_planejada": o.get("qtd_planejada", 0),
                "data_carga": data_carga,
                "cor": o.get("cor")
            })

        Ordem.objects.bulk_create(ordens_objs)
        
        ordens_numbers = [o.ordem for o in ordens_objs] 

        ordens_criadas = list(Ordem.objects.filter(
            ordem__in=ordens_numbers, 
            grupo_maquina=grupo_maquina
        ).order_by('ordem'))

        ordens_a_atualizar = []

        for ordem in ordens_criadas:
            
            caminho_qr_code = gerar_e_salvar_qrcode(request, ordem)
            
            if caminho_qr_code:
                ordem.caminho_relativo_qr_code = caminho_qr_code
                ordens_a_atualizar.append(ordem)

        if ordens_a_atualizar:
            Ordem.objects.bulk_update(ordens_a_atualizar, ['caminho_relativo_qr_code'])

        pecas_objs = [
            POM(
                ordem=ordem,
                peca=meta["peca_nome"],
                qtd_planejada=meta["qtd_planejada"],
                qtd_boa=0,
                qtd_morta=0
            ) for ordem, meta in zip(ordens_objs, ordens_metadata)
        ]

        POM.objects.bulk_create(pecas_objs)
    
        return {
            "message": "Ordens criadas com sucesso.",
            "ordens": [
                {
                    "id": ordem.id,
                    "data_carga": meta["data_carga"]#.strftime("%Y-%m-%d")
                } for ordem, meta in zip(ordens_objs, ordens_metadata)
            ]
        }

def processar_ordens_pintura(ordens_data, atualizacao_ordem=None, grupo_maquina="pintura"):

    if not ordens_data:
        return {"error": "Nenhuma ordem nova adicionada, caso tivesse ordem para ser excluída, foi excluída.", "status": 400}

    # Coletar datas únicas e validar
    try:
        formato_data = "%Y-%d-%m" if grupo_maquina == "montagem" or "solda" else "%Y-%m-%d"
        datas_requisicao = {
            datetime.strptime(o["data_carga"], "%Y-%m-%d").date()
            for o in ordens_data if o.get("data_carga")
        }
    except ValueError:
        return {"error": "Formato de data inválido! Use YYYY-MM-DD.", "status": 400}

    # Verifica datas já com carga
    datas_existentes = set(
        Ordem.objects.filter(data_carga__in=datas_requisicao, grupo_maquina=grupo_maquina)
        .values_list("data_carga", flat=True)
    )
    
    datas_bloqueadas = datas_existentes & datas_requisicao
    if not atualizacao_ordem and datas_bloqueadas:
        return {
            "error": f"Datas já com carga alocada: {', '.join(map(str, datas_bloqueadas))}",
            "status": 400
        }

    pecas_objs = []

    ultimo_numero = Ordem.objects.filter(grupo_maquina=grupo_maquina).aggregate(
        Max('ordem')
    )['ordem__max'] or 0

    with transaction.atomic():
        ordens_objs = []
        ordens_metadata = []

        for i, o in enumerate(ordens_data):
            data_carga = datetime.strptime(o["data_carga"], "%Y-%m-%d").date()
            peca_nome = o["peca_nome"]
            cor = o["cor"]

            # Tenta encontrar ordem existente com essa peça e data
            ordem_existente = Ordem.objects.filter(
                grupo_maquina=grupo_maquina,
                data_carga=data_carga,
                ordem_pecas_pintura__peca=peca_nome,
                cor=cor
            ).first()

            if ordem_existente:
                # Atualiza a qtd_planejada na peça vinculada
                POP.objects.filter(
                    ordem=ordem_existente,
                    peca=peca_nome
                ).update(qtd_planejada=o.get("qtd_planejada", 0))

                continue  # Não cria nova ordem, já atualizou

            # Cria nova ordem
            nova_ordem = Ordem(
                grupo_maquina=grupo_maquina,
                status_atual="aguardando_iniciar",
                obs=o.get("obs", ""),
                cor=o.get("cor"),
                data_criacao=now(),
                data_carga=data_carga,
                ordem=ultimo_numero + i + 1
            )

            nova_ordem.data_programacao = data_carga - timedelta(days=1)
            while nova_ordem.data_programacao.weekday() in [5, 6]:
                nova_ordem.data_programacao -= timedelta(days=1)

            ordens_objs.append(nova_ordem)
            ordens_metadata.append({
                "peca_nome": o["peca_nome"],
                "qtd_planejada": o.get("qtd_planejada", 0),
                "data_carga": data_carga,
                "cor": o.get("cor")
            })

        Ordem.objects.bulk_create(ordens_objs)

        pecas_objs = [
            POP(
                ordem=ordem,
                peca=meta["peca_nome"],
                qtd_planejada=meta["qtd_planejada"],
                qtd_boa=0,
                qtd_morta=0
            ) for ordem, meta in zip(ordens_objs, ordens_metadata)
        ]

        POP.objects.bulk_create(pecas_objs)

        return {
            "message": "Ordens criadas com sucesso.",
            "ordens": [
                {
                    "id": ordem.id,
                    "cor": meta["cor"],
                    "data_carga": meta["data_carga"].strftime("%Y-%m-%d")
                } for ordem, meta in zip(ordens_objs, ordens_metadata)
            ]
        }

def processar_ordens_solda(ordens_data, atualizacao_ordem=None, grupo_maquina='solda'):

    if not ordens_data:
        return {"error": "Nenhuma ordem nova adicionada, caso tivesse ordem para ser excluída, foi excluída.", "status": 400}

    # Coletar datas únicas e validar
    try:
        datas_requisicao = set()
        for o in ordens_data:
            data_carga = o.get("data_carga")
            if data_carga:
                if isinstance(data_carga, str):
                    # Se for string, converte usando strptime
                    data_obj = datetime.strptime(data_carga, "%Y-%m-%d").date()
                elif isinstance(data_carga, (datetime, date)):
                    # Se já for datetime ou date, converte para date (se necessário)
                    data_obj = data_carga.date() if isinstance(data_carga, datetime) else data_carga
                else:
                    raise ValueError(f"Tipo inválido de data: {type(data_carga)}")
                datas_requisicao.add(data_obj)
    except Exception as e:
        print(f"Erro ao processar datas: {e}")
        return {"error": "Formato de data inválido! Use YYYY-MM-DD.", "status": 400}

    # Verifica datas já com carga
    datas_existentes = set(
        Ordem.objects.filter(data_carga__in=datas_requisicao, grupo_maquina=grupo_maquina)
        .values_list("data_carga", flat=True)
    )

    datas_bloqueadas = datas_existentes & datas_requisicao
    if not atualizacao_ordem and datas_bloqueadas:
        return {
            "error": f"Datas já com carga alocada: {', '.join(map(str, datas_bloqueadas))}",
            "status": 400
        }

    # Verifica máquinas existentes
    maquinas_requisicao = {o.get("setor_conjunto") for o in ordens_data if o.get("setor_conjunto")}
    maquinas_existentes = set(Maquina.objects.filter(nome__in=maquinas_requisicao).values_list("nome", flat=True))
    maquinas_faltantes = maquinas_requisicao - maquinas_existentes
    if maquinas_faltantes:
        return {
            "error": f"Máquinas não cadastradas: {', '.join(maquinas_faltantes)}",
            "status": 400
        }

    # Criação em lote
    # Pega a última ordem atual no banco
    ultimo_numero = Ordem.objects.filter(grupo_maquina=grupo_maquina).aggregate(Max('ordem'))['ordem__max'] or 0

    with transaction.atomic():
        ordens_objs = []
        ordens_metadata = []

        for i, o in enumerate(ordens_data):
            # try:
            #     data_carga = datetime.strptime(o["data_carga"], "%Y-%d-%m").date()
            # except:
            # data_carga pode ser str, datetime ou date
            data_carga = o.get("data_carga")

            if isinstance(data_carga, str):
                data_carga = datetime.strptime(data_carga, "%Y-%m-%d").date()
            elif isinstance(data_carga, datetime):
                data_carga = data_carga.date()
            elif isinstance(data_carga, date):
                data_carga = data_carga
            else:
                raise ValueError(f"Tipo inválido para data_carga: {type(data_carga)}")

            nova_ordem = Ordem(
                grupo_maquina=grupo_maquina,
                status_atual="aguardando_iniciar",
                obs=o.get("obs", ""),
                cor=o.get("cor"),
                data_criacao=now(),
                data_carga=data_carga,
                ordem=ultimo_numero + i + 1  # atribui manualmente a ordem
            )

            print(o['setor_conjunto'])

            try:
                maquina = Maquina.objects.get(nome=o["setor_conjunto"], setor__nome='solda')
                nova_ordem.maquina = maquina
                # calcula data_programacao manualmente
                nova_ordem.data_programacao = data_carga - timedelta(days=3)
                while nova_ordem.data_programacao.weekday() in [5, 6]:
                    nova_ordem.data_programacao -= timedelta(days=1)
            except Maquina.DoesNotExist:
                return {"error": f"Máquina '{o['setor_conjunto']}' não cadastrada.", "status": 400}

            ordens_objs.append(nova_ordem)
            ordens_metadata.append({
                "peca_nome": o["peca_nome"],
                "qtd_planejada": o.get("qtd_planejada", 0),
                "data_carga": data_carga,
                "cor": o.get("cor")
            })

        Ordem.objects.bulk_create(ordens_objs)

        pecas_objs = [
            POS(
                ordem=ordem,
                peca=meta["peca_nome"],
                qtd_planejada=meta["qtd_planejada"],
                qtd_boa=0,
                qtd_morta=0
            ) for ordem, meta in zip(ordens_objs, ordens_metadata)
        ]

        POS.objects.bulk_create(pecas_objs)
    
        return {
            "message": "Ordens criadas com sucesso.",
            "ordens": [
                {
                    "id": ordem.id,
                    "data_carga": meta["data_carga"]#.strftime("%Y-%m-%d")
                } for ordem, meta in zip(ordens_objs, ordens_metadata)
            ]
        }

def imprimir_ordens_montagem(itens_agrupado_filtrado):#data_carga_str, ):

    CODIGOS_EXCLUIR = [
        '023590', '030679', '031517', '032470', '032546', '032531', '032681', '032731', '032871',
        '411528', '032637', '411267', '030499', '033053', '034015', '033884', '025118', '034316',
        '033601', '033387', '033400', '033750', '034375', '034306', '034467', '035939', '035945',
        '032705', '032992', '033121','033331','025782',
    ]

    impressas = 0
    ultima_celula = None
    # ultima_carga = None

    itens_agrupado_filtrado = itens_agrupado_filtrado.sort_values(['Célula','Datas'])
    
    for index, row in itens_agrupado_filtrado.iterrows():
        
        # retirando codigos_excluir da coluna peca
        # precisa ser like pois o codigo contem a descrição
        if any(codigo in str(row['Código']) for codigo in CODIGOS_EXCLUIR):
            continue
        
        celula_atual = row['Célula']
        carga_atual = row['Carga']  # ou pega do peca.ordem se houver campo específico
        data_carga = pd.to_datetime(row["Datas"], errors="coerce")
        data_carga_fmt = data_carga.strftime("%d/%m/%Y") if pd.notna(data_carga) else ""

        qtd = int(row.get("Qtde_total", 0))
        if qtd <= 0:
            continue

        # IMPRIME ETIQUETA DE TROCA DE CÉLULA
        if celula_atual != ultima_celula:
            print(f"indo para celula {celula_atual}")

            zpl_celula = f"""
^XA
^CI28
^PW560
^LL240
^LT0
^LH0,0

^FX CÉLULA
^FO10,35^A0N,32,32^FB540,1,0,C,0^FD{celula_atual}^FS

^XZ

""".lstrip()
            chamar_impressora_pecas_montagem(zpl_celula)
            ultima_celula = celula_atual

        for _ in range(qtd):
            zpl = f"""
^XA
^MMT
^PW560
^LL220
^LS0
^PR1,1,1
~SD14

^FO25,10
^A0N,28,28
^FB525,2,0,L,0
^FD{f'{row["Código"]}-{str(row["Peca"])[:80]}-{qtd}un.'}^FS

^FO25,80
^A0N,24,24
^FB525,1,0,L,0
^FDCelula: {celula_atual}^FS

^FO25,130
^A0N,24,24
^FB525,1,0,L,0
^FDCarga: {carga_atual}^FS

^FO25,180
^A0N,24,24
^FB525,1,0,L,0
^FDDt. Carga: {data_carga_fmt}^FS

^PQ1,0,0
^XZ

""".strip()
            chamar_impressora_pecas_montagem(zpl)
            impressas += 1

    if impressas == 0:
        return ({"error": "Nenhuma etiqueta a imprimir (qtd_planejada <= 0)."}, 422)

    return ({"message": "Impressão enviada com sucesso.", "total_etiquetas": impressas}, 200)

def imprimir_ordens_montagem_unitaria(ordem_id):
    # 1) buscar
    qs = (POM.objects
          .filter(
              ordem__id=ordem_id,
              qtd_boa=0,
          )
          .select_related('ordem'))

    if not qs.exists():
        return ({"error": "Nenhuma ordem encontrada para esse dia."}, 404)

    # 2) imprimir
    impressas = 0
    for peca in qs:
        qtd = int(peca.qtd_planejada or 0)
        if qtd <= 0:
            continue

        # Adiciona caminho_relativo_qr_code se não existir
        if not getattr(peca.ordem, 'caminho_relativo_qr_code', None):
            caminho_relativo = (
                reverse("montagem:apontamento_qrcode") + f"?ordem_id={peca.ordem.pk}&selecao_setor=pendente"
            )
            peca.ordem.caminho_relativo_qr_code = caminho_relativo
            peca.ordem.save(update_fields=['caminho_relativo_qr_code'])

        zpl = f"""
^XA

^MMT
^PW799
^LL0400
^LS0
^PR1,1,1
~SD14
^FO10,10
^A0N,40,40
^FB400,10,10,L,0
^FD{peca.peca[:80]}-{qtd}^FS

^FO10,280
^A0N,40,40
^FB400,10,10,L,0
^FDCarga: {peca.ordem.data_carga.strftime("%d/%m/%Y")}^FS

^FT500,330^BQN,2,7
^FDLA,{env('URI_QR_CODE')}{getattr(peca.ordem, 'caminho_relativo_qr_code', '')}^FS

^XZ
""".strip()

        chamar_impressora_pecas_montagem(zpl)
        impressas += qtd

    if impressas == 0:
        return ({"error": "Nenhuma etiqueta a imprimir (qtd_planejada <= 0)."}, 422)

    return ({"message": "Impressão enviada com sucesso.", "total_etiquetas": impressas}, 200)

def imprimir_ordens_pintura(data_carga, carga, itens_agrupados, pausa_s: float = 2.0):
    """
    data_carga: str | datetime | date  -> data que vai na etiqueta
    carga: str                          -> ex.: "Carga 01"
    itens_agrupados: pandas.DataFrame   -> precisa ter colunas: ["Código","Peca","Qtde_total"]
    pausa_s: float                      -> pausa entre etiquetas (segundos)
    """

    # Normaliza a data para datetime
    if isinstance(data_carga, str):
        try:
            data_carga_dt = datetime.fromisoformat(data_carga)
        except ValueError:
            # tenta outro formato comum
            data_carga_dt = datetime.strptime(data_carga, "%Y-%m-%d")
    elif isinstance(data_carga, date) and not isinstance(data_carga, datetime):
        data_carga_dt = datetime.combine(data_carga, datetime.min.time())
    else:
        data_carga_dt = data_carga

    total_impressoes = 0

    # Itera por cada linha agrupada
    for _, row in itens_agrupados.iterrows():
        codigo = str(row["Código"])
        peca = str(row["Peca"])[:80]  # limita a 80 chars para caber bem
        qtde_total = int(row.get("Qtde_total", 0) or 0)
        cor = str(row['cor'])
        if qtde_total <= 0:
            continue

        # Uma etiqueta por unidade: 1/Qtde_total, 2/Qtde_total, ...
        for i in range(1, qtde_total + 1):
            zpl = f"""
^XA
^MMT
^PW799
^LL0400
^LS0
^PR1,1,1
~SD14

^FO10,10
^A0N,40,40
^FB760,3,10,L,0
^FD{codigo} - {peca} - {i}/{qtde_total}^FS

^FO10,160
^A0N,40,40
^FB400,10,10,L,0
^FDCor: {cor}^FS

^FO10,220
^A0N,40,40
^FB760,2,10,L,0
^FDCarga: {carga}^FS

^FO10,280
^A0N,40,40
^FB760,2,10,L,0
^FDData carga: {data_carga_dt.strftime("%d/%m/%Y")}^FS

^XZ
""".lstrip()

            chamar_impressora_pecas_montagem(zpl)
            total_impressoes += 1
            time.sleep(pausa_s)  # controla o ritmo entre etiquetas

    return total_impressoes

# def send_raw_windows(zpl: str, printer_name: str) -> int:
#     data = zpl.encode("cp437", errors="replace")
#     h = win32print.OpenPrinter(printer_name)
#     try:
#         win32print.StartDocPrinter(h, 1, ("RAW ZPL", None, "RAW"))
#         win32print.StartPagePrinter(h)
#         written = win32print.WritePrinter(h, data)
#         win32print.EndPagePrinter(h)
#         win32print.EndDocPrinter(h)
#         return written or 0
#     finally:
#         win32print.ClosePrinter(h)

def enviar_para_impressao(zpl, str_celula, max_tentativas=3):
    """Envia uma etiqueta para impressão e aguarda confirmação do worker."""

    REDIS_URL = "redis://default:AWbmAbD4G2CfZPb3RxwuWQ4RfY7JOmxS@redis-19210.c262.us-east-1-3.ec2.redns.redis-cloud.com:19210"
    QUEUE_NAME = "print-zebra"
    r = redis.from_url(REDIS_URL)

    for tentativa in range(1, max_tentativas + 1):
        job_id = str(uuid.uuid4())
        resposta_queue = f"print-zebra:resposta:{job_id}"

        payload = {
            "job_id": job_id,
            "resposta_queue": resposta_queue,
            "zpl": zpl,
        }

        print(f"[{tentativa}/{max_tentativas}] Enviando job para impressão...")
        r.rpush(QUEUE_NAME, json.dumps(payload))

        # Espera até 10 segundos pela resposta do worker
        res = r.blpop(resposta_queue, timeout=0.2)

        if res:
            resposta = json.loads(res[1])
            status = resposta.get("status")
            print(f"[RESPOSTA] {status}")

            if status == "OK":
                print(f"✅ Impressão confirmada para célula {str_celula}")
                return True
            else:
                print(f"❌ Erro na impressão: {status}. Tentando novamente...")
        else:
            print("⚠️ Sem resposta do worker (timeout). Tentando reenviar...")

        time.sleep(0.2)  # espera antes de tentar novamente

    print(f"❌ Falha após {max_tentativas} tentativas. Impressora pode estar offline.")
    return False

def imprimir_ordens_pcp_qualidade(data_carga, carga, itens_agrupados, pausa_s: float = 1):
    """
    data_carga: str | datetime | date  -> data que vai na etiqueta
    carga: str                          -> ex.: "Carga 01"
    itens_agrupados: pandas.DataFrame   -> precisa ter colunas: ["Código","Peca","Qtde_total","Célula","cor"]
    pausa_s: float                      -> pausa entre etiquetas (segundos)
    """

    # Normaliza a data para datetime
    # if isinstance(data_carga, str):
    #     try:
    #         data_carga_dt = datetime.fromisoformat(data_carga)
    #     except ValueError:
    #         data_carga_dt = datetime.strptime(data_carga, "%Y-%m-%d")
    # elif isinstance(data_carga, date) and not isinstance(data_carga, datetime):
    #     data_carga_dt = datetime.combine(data_carga, datetime.min.time())
    # else:
    #     data_carga_dt = data_carga

    data_carga_dt = datetime.strptime(data_carga, "%Y-%m-%d")
    data_carga_str = datetime.strftime(data_carga_dt, "%d/%m/%Y")

    total_impressoes = 0
    ultima_str_celula = ''  # controla mudança de célula

    # Lê o logo e remove a âncora ^FO0,0 para reposicionar
    logo_gfa = open('logo.zpl').read().strip()
    logo_gfa_block = logo_gfa.replace("^FO0,0", "", 1)

    # Ordena as peças por célula (importante para o agrupamento)
    itens_agrupados = itens_agrupados.sort_values(by='Célula')

    reps = itens_agrupados['Qtde_total'].fillna(0).astype(int).clip(lower=0)

    # repete as linhas conforme a quantidade
    out = itens_agrupados.loc[itens_agrupados.index.repeat(reps)].copy()

    # cada linha passa a representar 1 unidade
    out['Qtde_total'] = 1

    # (opcional) numeração 1..N por linha original
    out["seq"] = out.groupby(level=0).cumcount() + 1
    out["seq_total"] = reps.loc[itens_agrupados.index.repeat(reps)].groupby(level=0).transform("size")
    out = out.reset_index(drop=True)

    printer = "ZDesigner ZD220-203dpi ZPL"

    timeout=60
    r = redis.from_url(REDIS_URL)
    ultima_cor = ""

    # out.to_csv("out1.csv")  
    # Itera por cada linha agrupada
    for _, row in out.iterrows():
        codigo = str(row["Código"])
        peca = str(row["Peca"])[:80]  # limita a 80 chars
        qtde_total = int(row.get("Qtde_total", 0) or 0)
        cor = str(row["cor"])
        str_celula = str(row["Célula"])

        # if qtde_total <= 0:
        #     continue

        if str_celula != ultima_str_celula or cor != ultima_cor:
            print(f"CELULA DIFERENTE: {str_celula[:11]} - {cor}")
            zpl = f"""
                ^XA
                ^CI28
                ^PW800
                ^LL320
                ^LT0
                ^LH0,0

                ^FX CÉLULA
                ^FO100,60^A0N,30,30^FB600,1,0,C,0^FDCélula: {str_celula}^FS

                ^FX COR
                ^FO100,160^A0N,30,30^FB600,1,0,C,0^FDCor: {cor}^FS

                ^FX DATA DA CARGA
                ^FO100,260^A0N,30,30^FB600,1,0,C,0^FDDATA CARGA: {data_carga_str}^FS

                ^XZ
            """.lstrip()

            sucesso = enviar_para_impressao(zpl, str_celula)
            if sucesso:
                total_impressoes += 1
                ultima_str_celula = str_celula
                ultima_cor = cor
                time.sleep(2)  # pausa antes do próximo job
            else:
                print(f"❌ Impressão da célula {str_celula} não confirmada.")
            # send_raw_windows(zpl_celula, printer)

            ultima_str_celula = str_celula

        # --- etiqueta normal da peça ---
        zpl = f"""
^XA
^CI28
^PW800
^LL320
^LT0
^LH0,0

^FX Logo centralizado
^FO220,1{logo_gfa_block}

^FX Código da peça e descrição
^FO50,110^A0N,28,28^FDCódigo: {codigo} - {peca}^FS

^FX Linha divisória vertical
^FO400,140^GB2,200,2^FS

^FX Coluna PCP (esquerda)
^FO50,160^A0N,36,36^FB350,1,0,C,0^FDPCP^FS
^FO60,210^A0N,30,30^FDCélula: {str_celula[:80]}^FS
^FO60,240^A0N,30,30^FDCor: {cor}^FS

^FX Coluna Qualidade (direita)
^FO420,160^A0N,36,36^FB330,1,0,C,0^FDQualidade^FS
^FO430,205^A0N,30,30^FDAprovado:^FS
^FO560,205^A0N,30,30^FDSim^FS
^FO610,205^GB20,20,20^FS
^FO650,205^A0N,30,30^FDNão^FS
^FO700,205^GB20,20,2^FS
^FO430,235^A0N,30,30^FDData: ___/___/___^FS
^FO430,270^A0N,30,30^FDInspetor: _____________^FS

^XZ
""".lstrip()

        sucesso = enviar_para_impressao(zpl, str_celula)
        print(codigo, peca)
        if sucesso:
            total_impressoes += 1
            ultima_str_celula = str_celula
            ultima_cor = cor
            time.sleep(2)  # pausa antes do próximo job
        else:
            print(f"❌ Impressão da célula {str_celula} não confirmada.")

        total_impressoes += 1
        # time.sleep(2)
        # send_raw_windows(zpl, printer)

    print(f"✅ Total de etiquetas impressas: {total_impressoes}")
    return total_impressoes

def gerar_e_salvar_qrcode(request, ordem):
    """
    Verifica as condições de uma ordem e, se atendidas,
    gera um QR Code e o salva no objeto.
    """

    # 74 = Chassi de Montagem, 37 = PLAT. TANQUE. CAÇAM.
    if (
        ordem.maquina
        and ordem.maquina.nome in ["CHASSI", "PLAT. TANQUE. CAÇAM."]
        and ordem.maquina.setor.nome == "montagem"
        and ordem.grupo_maquina == "montagem"
    ):

        # Gera a URL para o QR Code (agora temos certeza que o .pk existe)
        caminho_relativo = (
            reverse("montagem:apontamento_qrcode") + f"?ordem_id={ordem.pk}&selecao_setor=pendente"
        )

        url_completa = request.build_absolute_uri(caminho_relativo)

        print(url_completa)

        qr = qrcode.make(url_completa)
        qr_io = BytesIO()
        qr.save(qr_io, "PNG")

        file_name = f"ordem_{ordem.pk}_qrcode.png"

        ordem.qrcode.save(file_name, File(qr_io), save=False) 
        
        ordem.save(update_fields=['qrcode']) 
        
        return caminho_relativo
